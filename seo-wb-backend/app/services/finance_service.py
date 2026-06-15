import csv
import io
import json
from collections import defaultdict
from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

from google import genai
from google.genai import types
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import String, and_, func, or_
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.core.errors import AppError
from app.models.finance import (
    ExternalCost,
    FinanceAnalysisSnapshot,
    ProductFinanceSetting,
    SellerFinanceAutomationState,
    SellerFinanceSettings,
    WbFinanceReportRow,
    WbFinanceSyncState,
    WbFinancialDailySummary,
    WbFinancialMonthlySummary,
)
from app.models.seller import Seller
from app.models.wb_product import WbProduct, WbProductSyncState
from app.services.wb_base_client import WbNoData, WbRateLimitError, get_active_cooldowns
from app.services.wb_finance_client import WbFinanceClient


MONEY_QUANT = Decimal("0.0001")
STALE_SYNC_MINUTES = 15


def decimal_or_zero(value: Any) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def decimal_to_str(value: Decimal) -> str:
    return str(value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP))


def decimal_to_display_str(value: Decimal | None) -> str:
    if value is None:
        return ""
    normalized = decimal_or_zero(value)
    return str(int(normalized))


def decimal_from_input(value: Any, *, field_name: str) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        parsed = Decimal(str(value).strip())
    except Exception as exc:
        raise AppError("invalid_money_value", f"{field_name} must be a valid number.", 422) from exc
    if parsed < 0:
        raise AppError("invalid_money_value", f"{field_name} must be >= 0.", 422)
    return parsed


def dt_start(value: date) -> datetime:
    return datetime.combine(value, time.min, tzinfo=UTC)


def dt_end(value: date) -> datetime:
    return datetime.combine(value, time.max, tzinfo=UTC)


def validate_date_range(date_from: date, date_to: date) -> None:
    if date_to < date_from:
        raise AppError("invalid_date_range", "date_to must be greater than or equal to date_from.", 422)


def parse_dt(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError:
        return None


def parse_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    parsed = parse_dt(value)
    if parsed:
        return parsed.date()
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _product_setting_overlaps(
    db: Session,
    seller_id: int,
    product_id: int,
    effective_from: date,
    effective_to: date | None,
    exclude_id: int | None = None,
) -> bool:
    query = db.query(ProductFinanceSetting).filter(
        ProductFinanceSetting.seller_id == seller_id,
        ProductFinanceSetting.product_id == product_id,
    )
    if exclude_id is not None:
        query = query.filter(ProductFinanceSetting.id != exclude_id)
    for row in query.all():
        row_to = row.effective_to or date.max
        new_to = effective_to or date.max
        if row.effective_from <= new_to and effective_from <= row_to:
            return True
    return False


class FinanceSettingsService:
    def __init__(self, db: Session, seller: Seller) -> None:
        self._db = db
        self._seller = seller

    def get_seller_settings(self) -> SellerFinanceSettings:
        settings = self._db.query(SellerFinanceSettings).filter_by(seller_id=self._seller.id).one_or_none()
        if settings:
            return settings
        settings = SellerFinanceSettings(seller_id=self._seller.id)
        self._db.add(settings)
        self._db.commit()
        self._db.refresh(settings)
        return settings

    def update_seller_settings(self, payload: dict[str, Any]) -> SellerFinanceSettings:
        settings = self.get_seller_settings()
        for field in (
            "currency",
            "default_tax_mode",
            "tax_base",
        ):
            if field in payload and payload[field] is not None:
                setattr(settings, field, payload[field])
        for field in (
            "default_tax_rate",
            "default_packaging_cost",
            "default_labeling_cost",
            "default_shipping_to_warehouse_cost",
            "default_other_unit_cost",
        ):
            if field in payload and payload[field] is not None:
                value = decimal_or_zero(payload[field])
                if value < 0:
                    raise AppError("invalid_money_value", f"{field} must be >= 0.", 422)
                setattr(settings, field, value)
        self._db.commit()
        self._db.refresh(settings)
        return settings

    def list_product_settings(self) -> list[ProductFinanceSetting]:
        return (
            self._db.query(ProductFinanceSetting)
            .filter(ProductFinanceSetting.seller_id == self._seller.id)
            .order_by(ProductFinanceSetting.product_id.asc(), ProductFinanceSetting.effective_from.desc())
            .all()
        )

    def get_product_settings(self, product_id: int) -> list[ProductFinanceSetting]:
        return (
            self._db.query(ProductFinanceSetting)
            .filter(ProductFinanceSetting.seller_id == self._seller.id, ProductFinanceSetting.product_id == product_id)
            .order_by(ProductFinanceSetting.effective_from.desc())
            .all()
        )

    def list_product_settings_catalog(
        self,
        *,
        page: int = 1,
        per_page: int = 100,
        search: str | None = None,
        brands: list[str] | None = None,
        subjects: list[str] | None = None,
        only_missing: bool = False,
    ) -> dict[str, Any]:
        query = self._db.query(WbProduct).filter(WbProduct.seller_id == self._seller.id)
        if search:
            needle = f"%{search.strip()}%"
            query = query.filter(
                or_(
                    WbProduct.title.ilike(needle),
                    WbProduct.vendor_code.ilike(needle),
                    WbProduct.brand.ilike(needle),
                    WbProduct.subject_name.ilike(needle),
                    func.cast(WbProduct.nm_id, String).ilike(needle),
                )
            )
        if brands:
            query = query.filter(WbProduct.brand.in_(brands))
        if subjects:
            query = query.filter(WbProduct.subject_name.in_(subjects))

        product_ids_with_settings = {
            pid
            for (pid,) in self._db.query(ProductFinanceSetting.product_id)
            .filter(ProductFinanceSetting.seller_id == self._seller.id)
            .distinct()
            .all()
        }
        if only_missing and product_ids_with_settings:
            query = query.filter(~WbProduct.id.in_(product_ids_with_settings))

        total = query.count()
        products = (
            query.order_by(WbProduct.updated_at.desc(), WbProduct.id.desc())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )
        settings_map = self._latest_product_settings_map([product.id for product in products])
        items = [self._catalog_item_payload(product, settings_map.get(product.id)) for product in products]
        return {
            "items": items,
            "page": page,
            "perPage": per_page,
            "total": total,
            "facets": self._catalog_facets(),
        }

    def upsert_product_setting(self, product_id: int, payload: dict[str, Any]) -> ProductFinanceSetting:
        product = self._db.query(WbProduct).filter(WbProduct.id == product_id, WbProduct.seller_id == self._seller.id).one_or_none()
        if not product:
            raise AppError("product_not_found", "Product not found.", 404)
        effective_from = parse_date(payload.get("effective_from")) or date.today()
        effective_to = parse_date(payload.get("effective_to"))
        if effective_to and effective_to < effective_from:
            raise AppError("invalid_effective_range", "effective_to must be >= effective_from.", 422)

        setting_id = payload.get("id")
        setting = None
        if setting_id is not None:
            setting = (
                self._db.query(ProductFinanceSetting)
                .filter(ProductFinanceSetting.id == setting_id, ProductFinanceSetting.seller_id == self._seller.id)
                .one_or_none()
            )
        if setting is None:
            setting = ProductFinanceSetting(seller_id=self._seller.id, product_id=product_id)
            self._db.add(setting)
            self._db.flush()

        if _product_setting_overlaps(self._db, self._seller.id, product_id, effective_from, effective_to, setting.id):
            raise AppError("overlapping_effective_range", "Finance setting date ranges cannot overlap for the same product.", 422)

        for field in (
            "cost_price",
            "packaging_cost",
            "labeling_cost",
            "shipping_to_warehouse_cost",
            "other_unit_cost",
        ):
            value = decimal_or_zero(payload.get(field))
            if value < 0:
                raise AppError("invalid_money_value", f"{field} must be >= 0.", 422)
            setattr(setting, field, value)
        setting.cost_currency = payload.get("cost_currency") or "RUB"
        setting.tax_mode = payload.get("tax_mode")
        setting.tax_rate = decimal_or_zero(payload.get("tax_rate")) if payload.get("tax_rate") is not None else None
        setting.tax_base = payload.get("tax_base")
        setting.effective_from = effective_from
        setting.effective_to = effective_to
        setting.note = payload.get("note")
        self._db.commit()
        self._db.refresh(setting)
        return setting

    def bulk_upsert_product_cost_prices(self, items: list[dict[str, Any]]) -> list[ProductFinanceSetting]:
        if not items:
            return []

        product_ids = [int(item["product_id"]) for item in items]
        products = (
            self._db.query(WbProduct)
            .filter(WbProduct.seller_id == self._seller.id, WbProduct.id.in_(product_ids))
            .all()
        )
        products_by_id = {product.id: product for product in products}
        missing_product_ids = [product_id for product_id in product_ids if product_id not in products_by_id]
        if missing_product_ids:
            raise AppError("product_not_found", f"Products not found: {', '.join(str(value) for value in missing_product_ids[:10])}", 404)

        settings_map = self._latest_product_settings_map(product_ids)
        today = date.today()
        updated_rows: list[ProductFinanceSetting] = []

        for item in items:
            product_id = int(item["product_id"])
            cost_price = decimal_from_input(item.get("cost_price"), field_name="cost_price")

            setting = settings_map.get(product_id)
            if setting is None:
                setting = ProductFinanceSetting(
                    seller_id=self._seller.id,
                    product_id=product_id,
                    cost_currency="RUB",
                    cost_price=Decimal("0"),
                    packaging_cost=Decimal("0"),
                    labeling_cost=Decimal("0"),
                    shipping_to_warehouse_cost=Decimal("0"),
                    other_unit_cost=Decimal("0"),
                    effective_from=today,
                )
                self._db.add(setting)
                settings_map[product_id] = setting

            setting.cost_price = cost_price
            updated_rows.append(setting)

        self._db.commit()
        for row in updated_rows:
            self._db.refresh(row)
        return updated_rows

    def list_missing_finance_settings(self) -> list[WbProduct]:
        products = self._db.query(WbProduct).filter(WbProduct.seller_id == self._seller.id).all()
        product_ids = {row.product_id for row in self._db.query(ProductFinanceSetting.product_id).filter(ProductFinanceSetting.seller_id == self._seller.id)}
        return [product for product in products if product.id not in product_ids]

    def _latest_product_settings_map(self, product_ids: list[int]) -> dict[int, ProductFinanceSetting]:
        if not product_ids:
            return {}
        rows = (
            self._db.query(ProductFinanceSetting)
            .filter(ProductFinanceSetting.seller_id == self._seller.id, ProductFinanceSetting.product_id.in_(product_ids))
            .order_by(ProductFinanceSetting.product_id.asc(), ProductFinanceSetting.effective_from.desc(), ProductFinanceSetting.id.desc())
            .all()
        )
        out: dict[int, ProductFinanceSetting] = {}
        for row in rows:
            if row.product_id not in out:
                out[row.product_id] = row
        return out

    def _resolve_product_from_import_row(self, row: dict[str, Any]) -> WbProduct:
        product_id_raw = row.get("productId")
        if product_id_raw not in (None, ""):
            try:
                product_id = int(product_id_raw)
            except Exception as exc:
                raise AppError("invalid_product_id", "productId must be an integer.", 422) from exc
            product = (
                self._db.query(WbProduct)
                .filter(WbProduct.id == product_id, WbProduct.seller_id == self._seller.id)
                .one_or_none()
            )
            if product is None:
                raise AppError("product_not_found", f"Product not found for productId={product_id}.", 404)
            return product

        nm_id_raw = row.get("nmId")
        if nm_id_raw in (None, ""):
            raise AppError("missing_product_key", "Each row must include productId or nmId.", 422)
        try:
            nm_id = int(str(nm_id_raw).strip())
        except Exception as exc:
            raise AppError("invalid_nm_id", "nmId must be an integer.", 422) from exc
        product = (
            self._db.query(WbProduct)
            .filter(WbProduct.seller_id == self._seller.id, WbProduct.nm_id == nm_id)
            .one_or_none()
        )
        if product is None:
            raise AppError("product_not_found", f"Product not found for nmId={nm_id}.", 404)
        return product

    def _catalog_item_payload(self, product: WbProduct, setting: ProductFinanceSetting | None) -> dict[str, Any]:
        return {
            "productId": product.id,
            "nmId": product.nm_id,
            "vendorCode": product.vendor_code,
            "title": product.title,
            "subjectName": product.subject_name,
            "brand": product.brand,
            "photoSquareUrl": product.photo_square_url,
            "photoBigUrl": product.photo_big_url,
            "hasCostSettings": setting is not None,
            "settingId": setting.id if setting else None,
            "costPrice": decimal_to_str(setting.cost_price) if setting else None,
            "costCurrency": setting.cost_currency if setting else None,
            "packagingCost": decimal_to_str(setting.packaging_cost) if setting else None,
            "labelingCost": decimal_to_str(setting.labeling_cost) if setting else None,
            "shippingToWarehouseCost": decimal_to_str(setting.shipping_to_warehouse_cost) if setting else None,
            "otherUnitCost": decimal_to_str(setting.other_unit_cost) if setting else None,
            "taxMode": setting.tax_mode if setting else None,
            "taxRate": decimal_to_str(setting.tax_rate) if setting and setting.tax_rate is not None else None,
            "taxBase": setting.tax_base if setting else None,
            "effectiveFrom": setting.effective_from.isoformat() if setting else None,
            "effectiveTo": setting.effective_to.isoformat() if setting and setting.effective_to else None,
            "note": setting.note if setting else None,
        }

    def _catalog_facets(self) -> dict[str, list[str]]:
        brands = [
            row[0]
            for row in self._db.query(WbProduct.brand)
            .filter(WbProduct.seller_id == self._seller.id, WbProduct.brand.is_not(None), WbProduct.brand != "")
            .distinct()
            .order_by(WbProduct.brand.asc())
            .all()
        ]
        subjects = [
            row[0]
            for row in self._db.query(WbProduct.subject_name)
            .filter(WbProduct.seller_id == self._seller.id, WbProduct.subject_name.is_not(None), WbProduct.subject_name != "")
            .distinct()
            .order_by(WbProduct.subject_name.asc())
            .all()
        ]
        return {"brands": brands, "subjects": subjects}

    def list_external_costs(self) -> list[ExternalCost]:
        return self._db.query(ExternalCost).filter(ExternalCost.seller_id == self._seller.id).order_by(ExternalCost.cost_date.desc(), ExternalCost.id.desc()).all()

    def create_external_cost(self, payload: dict[str, Any]) -> ExternalCost:
        amount = decimal_or_zero(payload.get("amount"))
        if amount < 0:
            raise AppError("invalid_money_value", "amount must be >= 0.", 422)
        row = ExternalCost(
            seller_id=self._seller.id,
            cost_date=parse_date(payload.get("cost_date")) or date.today(),
            period_from=parse_date(payload.get("period_from")),
            period_to=parse_date(payload.get("period_to")),
            cost_type=str(payload.get("cost_type") or "other"),
            amount=amount,
            currency=payload.get("currency") or "RUB",
            allocation_method=payload.get("allocation_method") or "BY_REVENUE",
            product_id=payload.get("product_id"),
            note=payload.get("note"),
        )
        self._db.add(row)
        self._db.commit()
        self._db.refresh(row)
        return row

    def update_external_cost(self, cost_id: int, payload: dict[str, Any]) -> ExternalCost:
        row = self._db.query(ExternalCost).filter(ExternalCost.id == cost_id, ExternalCost.seller_id == self._seller.id).one_or_none()
        if not row:
            raise AppError("external_cost_not_found", "External cost not found.", 404)
        for field in ("cost_type", "currency", "allocation_method", "note"):
            if field in payload and payload[field] is not None:
                setattr(row, field, payload[field])
        for field in ("cost_date", "period_from", "period_to"):
            if field in payload:
                setattr(row, field, parse_date(payload[field]) if payload[field] else None)
        if "product_id" in payload:
            row.product_id = payload["product_id"]
        if "amount" in payload:
            amount = decimal_or_zero(payload["amount"])
            if amount < 0:
                raise AppError("invalid_money_value", "amount must be >= 0.", 422)
            row.amount = amount
        self._db.commit()
        self._db.refresh(row)
        return row

    def delete_external_cost(self, cost_id: int) -> None:
        row = self._db.query(ExternalCost).filter(ExternalCost.id == cost_id, ExternalCost.seller_id == self._seller.id).one_or_none()
        if not row:
            raise AppError("external_cost_not_found", "External cost not found.", 404)
        self._db.delete(row)
        self._db.commit()

    def export_product_settings_template_xlsx(self, *, include_values: bool) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "product-costs"
        editable_fill = PatternFill(fill_type="solid", fgColor="FFF59D")
        sheet.append(
            [
                "productId",
                "nmId",
                "vendorCode",
                "title",
                "costPrice",
            ]
        )
        products = self._db.query(WbProduct).filter(WbProduct.seller_id == self._seller.id).order_by(WbProduct.id.asc()).all()
        settings_map = self._latest_product_settings_map([product.id for product in products])
        for product in products:
            setting = settings_map.get(product.id)
            sheet.append(
                [
                    product.id,
                    product.nm_id,
                    product.vendor_code or "",
                    product.title or "",
                    decimal_to_display_str(setting.cost_price) if include_values and setting else "",
                ]
            )
        cost_price_col = 5
        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row=row_index, column=cost_price_col).fill = editable_fill
        for column_cells in sheet.columns:
            values = [len(str(cell.value or "")) for cell in column_cells]
            width = max(values, default=10) + 3
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width, 12), 40)
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def import_product_settings_file(self, filename: str, content: bytes) -> dict[str, Any]:
        lower_name = filename.lower()
        if lower_name.endswith(".csv"):
            decoded = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(decoded))
            return self._import_product_settings_rows(reader)
        if lower_name.endswith(".xlsx"):
            workbook = load_workbook(filename=BytesIO(content), data_only=True)
            sheet = workbook.active
            headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            rows = [dict(zip(headers, [cell.value for cell in row], strict=False)) for row in sheet.iter_rows(min_row=2)]
            return self._import_product_settings_rows(rows)
        raise AppError("unsupported_import_file", "Only .xlsx and .csv imports are supported.", 422)

    def _import_product_settings_rows(self, rows: Any) -> dict[str, Any]:
        imported = 0
        errors: list[dict[str, Any]] = []
        required_columns = {"costPrice"}
        first_row = next(iter(rows), None)
        if first_row is None:
            return {"imported": 0, "errors": []}
        if not required_columns.issubset(set(first_row.keys())):
            raise AppError("invalid_import_columns", "Import file is missing required column: costPrice.", 422)
        all_rows = [first_row, *list(rows)] if not isinstance(rows, list) else rows
        payload_items: list[dict[str, Any]] = []
        for index, row in enumerate(all_rows, start=2):
            try:
                product = self._resolve_product_from_import_row(row)
                cost_price = decimal_from_input(row.get("costPrice"), field_name="costPrice")
                payload_items.append({"product_id": product.id, "cost_price": str(cost_price)})
            except Exception as exc:
                errors.append({"row": index, "error": str(exc)})
        if payload_items:
            self.bulk_upsert_product_cost_prices(payload_items)
            imported = len(payload_items)
        return {"imported": imported, "errors": errors}


class FinanceSyncService:
    def __init__(self, db: Session, settings: Settings, seller: Seller, client: WbFinanceClient) -> None:
        self._db = db
        self._settings = settings
        self._seller = seller
        self._client = client

    async def sync(self, *, date_from: date, date_to: date, period: str, force: bool = False, max_pages: int | None = None) -> dict[str, Any]:
        validate_date_range(date_from, date_to)
        state = (
            self._db.query(WbFinanceSyncState)
            .filter(
                WbFinanceSyncState.seller_id == self._seller.id,
                WbFinanceSyncState.date_from == date_from,
                WbFinanceSyncState.date_to == date_to,
                WbFinanceSyncState.period == period,
            )
            .one_or_none()
        )
        if state is None:
            state = WbFinanceSyncState(seller_id=self._seller.id, date_from=date_from, date_to=date_to, period=period)
            self._db.add(state)
            self._db.commit()
            self._db.refresh(state)
        elif state.status == "running" and self.is_stale_state(state):
            state.status = "interrupted"
            state.last_error = "stale running finance sync recovered"
            state.finished_at = datetime.now(UTC)
            self._db.commit()
        if force:
            state.last_rrd_id = 0
            state.total_rows = 0

        state.status = "running"
        state.last_error = None
        state.started_at = datetime.now(UTC)
        state.finished_at = None
        self._db.commit()
        state_id = state.id

        last_rrd_id = state.last_rrd_id
        inserted = 0
        page_count = 0
        try:
            while True:
                try:
                    rows = await self._client.get_sales_reports_detailed_by_period(
                        date_from=dt_start(date_from),
                        date_to=dt_end(date_to),
                        period=period,
                        rrd_id=last_rrd_id,
                    )
                except WbNoData:
                    state.status = "completed"
                    state.finished_at = datetime.now(UTC)
                    self._db.commit()
                    break
                except WbRateLimitError as exc:
                    state.status = "rate_limited"
                    state.last_error = f"{exc.message} retry_after_seconds={exc.details.get('retry_after_seconds')}"
                    state.finished_at = datetime.now(UTC)
                    self._db.commit()
                    raise

                if not rows:
                    state.status = "completed"
                    state.finished_at = datetime.now(UTC)
                    self._db.commit()
                    break

                for raw_row in rows:
                    if raw_row.get("nmId") in (None, ""):
                        self._delete_report_row_by_rrd_id(raw_row)
                        continue
                    self._upsert_report_row(raw_row)
                    inserted += 1
                last_rrd_id = max(int(row.get("rrdId") or 0) for row in rows)
                state.last_rrd_id = last_rrd_id
                state.total_rows = (
                    self._db.query(func.count(WbFinanceReportRow.id))
                    .filter(WbFinanceReportRow.seller_id == self._seller.id, WbFinanceReportRow.date_from == date_from, WbFinanceReportRow.date_to == date_to)
                    .scalar()
                    or 0
                )
                self._db.commit()
                page_count += 1
                if max_pages is not None and page_count >= max_pages:
                    break

            return {"status": state.status, "rowsInserted": inserted, "lastRrdId": state.last_rrd_id}
        except Exception as exc:
            self._db.rollback()
            failed_state = self._db.get(WbFinanceSyncState, state_id)
            if failed_state is not None and failed_state.status == "running":
                failed_state.status = "failed"
                failed_state.last_error = str(exc)[:1000]
                failed_state.finished_at = datetime.now(UTC)
                self._db.commit()
            raise

    def _find_product_id(self, raw_row: dict[str, Any]) -> int | None:
        nm_id = raw_row.get("nmId")
        if nm_id is not None:
            product = self._db.query(WbProduct).filter(WbProduct.seller_id == self._seller.id, WbProduct.nm_id == int(nm_id)).one_or_none()
            if product:
                return product.id
        return None

    def _delete_report_row_by_rrd_id(self, raw_row: dict[str, Any]) -> None:
        rrd_id = int(raw_row.get("rrdId") or 0)
        if not rrd_id:
            return
        row = self._db.query(WbFinanceReportRow).filter(WbFinanceReportRow.seller_id == self._seller.id, WbFinanceReportRow.rrd_id == rrd_id).one_or_none()
        if row is None:
            return
        self._db.delete(row)
        self._db.commit()

    def _upsert_report_row(self, raw_row: dict[str, Any]) -> WbFinanceReportRow:
        rrd_id = int(raw_row.get("rrdId") or 0)
        row = self._db.query(WbFinanceReportRow).filter(WbFinanceReportRow.seller_id == self._seller.id, WbFinanceReportRow.rrd_id == rrd_id).one_or_none()
        if row is None:
            row = WbFinanceReportRow(seller_id=self._seller.id, rrd_id=rrd_id, raw_data={})
            self._db.add(row)
        for attr, key in (
            ("report_id", "reportId"),
            ("nm_id", "nmId"),
            ("brand_name", "brandName"),
            ("vendor_code", "vendorCode"),
            ("title", "title"),
            ("subject_name", "subjectName"),
            ("tech_size", "techSize"),
            ("sku", "sku"),
            ("doc_type_name", "docTypeName"),
            ("seller_oper_name", "sellerOperName"),
            ("office_name", "officeName"),
            ("payment_processing", "paymentProcessing"),
            ("acquiring_bank", "acquiringBank"),
            ("order_id", "orderId"),
            ("order_uid", "orderUid"),
            ("srid", "srid"),
            ("kiz", "kiz"),
            ("delivery_method", "deliveryMethod"),
            ("currency", "currency"),
            ("report_type", "reportType"),
            ("shk_id", "shkId"),
        ):
            setattr(row, attr, raw_row.get(key))
        row.date_from = parse_date(raw_row.get("dateFrom"))
        row.date_to = parse_date(raw_row.get("dateTo"))
        row.create_date = parse_date(raw_row.get("createDate"))
        row.order_dt = parse_dt(raw_row.get("orderDt"))
        row.sale_dt = parse_dt(raw_row.get("saleDt"))
        row.rr_date = parse_date(raw_row.get("rrDate"))
        row.quantity = int(raw_row.get("quantity") or 0)
        row.delivery_amount = int(raw_row.get("deliveryAmount") or 0)
        row.return_amount = int(raw_row.get("returnAmount") or 0)
        row.is_b2b = raw_row.get("isB2b")
        for attr, key in (
            ("retail_price", "retailPrice"),
            ("retail_amount", "retailAmount"),
            ("retail_price_with_disc", "retailPriceWithDisc"),
            ("sale_percent", "salePercent"),
            ("commission_percent", "commissionPercent"),
            ("delivery_service", "deliveryService"),
            ("ppvz_sales_commission", "ppvzSalesCommission"),
            ("for_pay", "forPay"),
            ("acquiring_fee", "acquiringFee"),
            ("acquiring_percent", "acquiringPercent"),
            ("penalty", "penalty"),
            ("additional_payment", "additionalPayment"),
            ("rebill_logistic_cost", "rebillLogisticCost"),
            ("paid_storage", "paidStorage"),
            ("deduction", "deduction"),
            ("paid_acceptance", "paidAcceptance"),
            ("cashback_amount", "cashbackAmount"),
            ("cashback_discount", "cashbackDiscount"),
            ("cashback_commission_change", "cashbackCommissionChange"),
            ("agency_vat", "agencyVat"),
        ):
            setattr(row, attr, decimal_or_zero(raw_row.get(key)))
        row.product_id = self._find_product_id(raw_row)
        row.raw_data = raw_row
        self._db.commit()
        self._db.refresh(row)
        return row

    @staticmethod
    def is_stale_state(state: WbFinanceSyncState) -> bool:
        if state.status != "running" or not state.updated_at:
            return False
        return (datetime.now(UTC) - state.updated_at.astimezone(UTC)).total_seconds() > STALE_SYNC_MINUTES * 60


class ProfitCalculationService:
    def __init__(self, db: Session, seller: Seller) -> None:
        self._db = db
        self._seller = seller

    def _seller_defaults(self) -> SellerFinanceSettings:
        return FinanceSettingsService(self._db, self._seller).get_seller_settings()

    def product_cost_for_date(self, product_id: int | None, target_date: date | None) -> tuple[Decimal, dict[str, Any], bool]:
        defaults = self._seller_defaults()
        if product_id is None or target_date is None:
            return Decimal("0"), {
                "costPrice": decimal_to_str(Decimal("0")),
                "packagingCost": decimal_to_str(defaults.default_packaging_cost),
                "labelingCost": decimal_to_str(defaults.default_labeling_cost),
                "shippingToWarehouseCost": decimal_to_str(defaults.default_shipping_to_warehouse_cost),
                "otherUnitCost": decimal_to_str(defaults.default_other_unit_cost),
            }, False
        setting = (
            self._db.query(ProductFinanceSetting)
            .filter(
                ProductFinanceSetting.seller_id == self._seller.id,
                ProductFinanceSetting.product_id == product_id,
                ProductFinanceSetting.effective_from <= target_date,
                or_(ProductFinanceSetting.effective_to.is_(None), ProductFinanceSetting.effective_to >= target_date),
            )
            .order_by(ProductFinanceSetting.effective_from.desc())
            .first()
        )
        if not setting:
            return Decimal("0"), {
                "costPrice": decimal_to_str(Decimal("0")),
                "packagingCost": decimal_to_str(defaults.default_packaging_cost),
                "labelingCost": decimal_to_str(defaults.default_labeling_cost),
                "shippingToWarehouseCost": decimal_to_str(defaults.default_shipping_to_warehouse_cost),
                "otherUnitCost": decimal_to_str(defaults.default_other_unit_cost),
            }, False
        total = setting.cost_price + setting.packaging_cost + setting.labeling_cost + setting.shipping_to_warehouse_cost + setting.other_unit_cost
        return total, {
            "costPrice": decimal_to_str(setting.cost_price),
            "packagingCost": decimal_to_str(setting.packaging_cost),
            "labelingCost": decimal_to_str(setting.labeling_cost),
            "shippingToWarehouseCost": decimal_to_str(setting.shipping_to_warehouse_cost),
            "otherUnitCost": decimal_to_str(setting.other_unit_cost),
            "taxMode": setting.tax_mode or defaults.default_tax_mode,
            "taxRate": decimal_to_str(setting.tax_rate or defaults.default_tax_rate),
            "taxBase": setting.tax_base or defaults.tax_base,
        }, True

    def tax_amount(self, *, gross_revenue: Decimal, profit_before_tax: Decimal, product_meta: dict[str, Any]) -> Decimal:
        defaults = self._seller_defaults()
        tax_mode = product_meta.get("taxMode") or defaults.default_tax_mode
        tax_rate = decimal_or_zero(product_meta.get("taxRate")) or defaults.default_tax_rate
        tax_base = product_meta.get("taxBase") or defaults.tax_base
        if tax_mode == "NONE" or tax_rate <= 0:
            return Decimal("0")
        if tax_mode == "USN_INCOME" or tax_base == "REVENUE":
            return max(gross_revenue, Decimal("0")) * tax_rate
        return max(profit_before_tax, Decimal("0")) * tax_rate


class FinanceAggregationService:
    def __init__(self, db: Session, seller: Seller) -> None:
        self._db = db
        self._seller = seller
        self._profit_service = ProfitCalculationService(db, seller)

    def _query_rows(self, date_from: date, date_to: date):
        return (
            self._db.query(WbFinanceReportRow)
            .filter(
                WbFinanceReportRow.seller_id == self._seller.id,
                or_(
                    and_(WbFinanceReportRow.rr_date.is_not(None), WbFinanceReportRow.rr_date >= date_from, WbFinanceReportRow.rr_date <= date_to),
                    and_(WbFinanceReportRow.sale_dt.is_not(None), func.date(WbFinanceReportRow.sale_dt) >= date_from, func.date(WbFinanceReportRow.sale_dt) <= date_to),
                ),
            )
        )

    def _external_cost_allocations(self, date_from: date, date_to: date, product_totals: dict[int | None, dict[str, Decimal]]) -> dict[int | None, Decimal]:
        costs = (
            self._db.query(ExternalCost)
            .filter(
                ExternalCost.seller_id == self._seller.id,
                or_(
                    and_(ExternalCost.period_from.is_not(None), ExternalCost.period_to.is_not(None), ExternalCost.period_from <= date_to, ExternalCost.period_to >= date_from),
                    and_(ExternalCost.period_from.is_(None), ExternalCost.cost_date >= date_from, ExternalCost.cost_date <= date_to),
                ),
            )
            .all()
        )
        allocations: dict[int | None, Decimal] = defaultdict(lambda: Decimal("0"))
        revenue_total = sum((item["grossRevenue"] for item in product_totals.values()), Decimal("0"))
        quantity_total = sum((item["quantity"] for item in product_totals.values()), Decimal("0"))
        product_count = Decimal(str(max(len([k for k in product_totals if k is not None]), 1)))
        for cost in costs:
            amount = decimal_or_zero(cost.amount)
            if cost.allocation_method == "DIRECT_PRODUCT" and cost.product_id:
                allocations[cost.product_id] += amount
            elif cost.allocation_method == "BY_SOLD_QUANTITY" and quantity_total > 0:
                for product_id, item in product_totals.items():
                    allocations[product_id] += amount * (item["quantity"] / quantity_total)
            elif cost.allocation_method == "EQUAL_BY_PRODUCT":
                share = amount / product_count
                for product_id in product_totals:
                    if product_id is not None:
                        allocations[product_id] += share
            elif cost.allocation_method == "MANUAL_NONE":
                allocations[None] += amount
            else:
                if revenue_total <= 0:
                    allocations[None] += amount
                else:
                    for product_id, item in product_totals.items():
                        allocations[product_id] += amount * (item["grossRevenue"] / revenue_total)
        return allocations

    def _product_rollup(self, rows: list[WbFinanceReportRow], date_from: date, date_to: date) -> tuple[list[dict[str, Any]], dict[int | None, dict[str, Decimal]]]:
        grouped: dict[int | None, list[WbFinanceReportRow]] = defaultdict(list)
        for row in rows:
            grouped[row.product_id].append(row)

        product_totals: dict[int | None, dict[str, Decimal]] = {}
        for product_id, items in grouped.items():
            product_totals[product_id] = {
                "grossRevenue": sum((decimal_or_zero(item.retail_amount) for item in items), Decimal("0")),
                "forPay": sum((decimal_or_zero(item.for_pay) for item in items), Decimal("0")),
                "quantity": sum((Decimal(item.quantity) for item in items), Decimal("0")),
            }
        allocations = self._external_cost_allocations(date_from, date_to, product_totals)

        results = []
        for product_id, items in grouped.items():
            product = self._db.get(WbProduct, product_id) if product_id else None
            gross_revenue = sum((decimal_or_zero(item.retail_amount) for item in items), Decimal("0"))
            for_pay = sum((decimal_or_zero(item.for_pay) for item in items), Decimal("0"))
            wb_costs = sum(
                (
                    decimal_or_zero(item.delivery_service)
                    + decimal_or_zero(item.ppvz_sales_commission)
                    + decimal_or_zero(item.acquiring_fee)
                    + decimal_or_zero(item.penalty)
                    + decimal_or_zero(item.deduction)
                    + decimal_or_zero(item.paid_storage)
                    + decimal_or_zero(item.paid_acceptance)
                    + decimal_or_zero(item.rebill_logistic_cost)
                    - decimal_or_zero(item.additional_payment)
                )
                for item in items
            )
            quantity = sum((Decimal(item.quantity) for item in items), Decimal("0"))
            reference_date = next((item.rr_date or (item.sale_dt.date() if item.sale_dt else None) for item in items), None)
            unit_cost, product_meta, has_cost = self._profit_service.product_cost_for_date(product_id, reference_date)
            cogs = unit_cost * quantity
            external_allocated_costs = allocations.get(product_id, Decimal("0"))
            profit_before_tax = for_pay - cogs - external_allocated_costs
            tax_amount = self._profit_service.tax_amount(gross_revenue=gross_revenue, profit_before_tax=profit_before_tax, product_meta=product_meta)
            profit_after_tax = profit_before_tax - tax_amount
            profit_margin = (profit_after_tax / gross_revenue * Decimal("100")) if gross_revenue > 0 else Decimal("0")
            results.append(
                {
                    "productId": product_id,
                    "nmId": product.nm_id if product else next((item.nm_id for item in items if item.nm_id), None),
                    "vendorCode": product.vendor_code if product else next((item.vendor_code for item in items if item.vendor_code), None),
                    "title": product.title if product else next((item.title for item in items if item.title), None),
                    "quantity": str(quantity),
                    "grossRevenue": decimal_to_str(gross_revenue),
                    "forPay": decimal_to_str(for_pay),
                    "wbCosts": decimal_to_str(wb_costs),
                    "cogs": decimal_to_str(cogs),
                    "externalAllocatedCosts": decimal_to_str(external_allocated_costs),
                    "profitBeforeTax": decimal_to_str(profit_before_tax),
                    "taxAmount": decimal_to_str(tax_amount),
                    "profitAfterTax": decimal_to_str(profit_after_tax),
                    "profitMargin": decimal_to_str(profit_margin),
                    "hasCostSettings": has_cost,
                    "costMeta": product_meta,
                }
            )
        return results, product_totals

    def summary(self, date_from: date, date_to: date) -> dict[str, Any]:
        validate_date_range(date_from, date_to)
        
        summaries = (
            self._db.query(WbFinancialDailySummary)
            .filter(
                WbFinancialDailySummary.seller_id == self._seller.id,
                WbFinancialDailySummary.summary_date >= date_from,
                WbFinancialDailySummary.summary_date <= date_to,
            )
            .all()
        )
        
        if not summaries:
            return {
                "period": {"dateFrom": date_from.isoformat(), "dateTo": date_to.isoformat()},
                "grossRevenue": "0.0000",
                "forPay": "0.0000",
                "wbCosts": "0.0000",
                "cogs": "0.0000",
                "externalAllocatedCosts": "0.0000",
                "profitBeforeTax": "0.0000",
                "taxAmount": "0.0000",
                "profitAfterTax": "0.0000",
                "profitMargin": "0.0000",
                "costCompletenessPercent": "0.0000",
                "rowsCount": 0,
                "productsCount": 0,
            }
            
        gross_revenue = sum((s.gross_revenue for s in summaries), Decimal("0"))
        for_pay = sum((s.for_pay for s in summaries), Decimal("0"))
        wb_costs = sum((s.wb_costs for s in summaries), Decimal("0"))
        cogs = sum((s.cogs for s in summaries), Decimal("0"))
        tax_amount = sum((s.tax_amount for s in summaries), Decimal("0"))
        profit_before_tax = sum((s.profit_before_tax for s in summaries), Decimal("0"))
        profit_after_tax = sum((s.profit_after_tax for s in summaries), Decimal("0"))
        rows_count = sum((s.raw_row_count for s in summaries), 0)
        
        costs = (
            self._db.query(ExternalCost)
            .filter(
                ExternalCost.seller_id == self._seller.id,
                or_(
                    and_(ExternalCost.period_from.is_not(None), ExternalCost.period_to.is_not(None), ExternalCost.period_from <= date_to, ExternalCost.period_to >= date_from),
                    and_(ExternalCost.period_from.is_(None), ExternalCost.cost_date >= date_from, ExternalCost.cost_date <= date_to),
                ),
            )
            .all()
        )
        external_costs = sum((decimal_or_zero(c.amount) for c in costs), Decimal("0"))
        
        profit_before_tax -= external_costs
        profit_after_tax -= external_costs
        
        products_count = self._db.query(func.count(WbProduct.id)).filter(WbProduct.seller_id == self._seller.id).scalar() or 0
        complete_count = self._db.query(func.count(ProductFinanceSetting.id)).filter(ProductFinanceSetting.seller_id == self._seller.id).scalar() or 0
        cost_completeness = Decimal("0") if products_count == 0 else Decimal(complete_count) / Decimal(products_count) * Decimal("100")
        
        return {
            "period": {"dateFrom": date_from.isoformat(), "dateTo": date_to.isoformat()},
            "grossRevenue": decimal_to_str(gross_revenue),
            "forPay": decimal_to_str(for_pay),
            "wbCosts": decimal_to_str(wb_costs),
            "cogs": decimal_to_str(cogs),
            "externalAllocatedCosts": decimal_to_str(external_costs),
            "profitBeforeTax": decimal_to_str(profit_before_tax),
            "taxAmount": decimal_to_str(tax_amount),
            "profitAfterTax": decimal_to_str(profit_after_tax),
            "profitMargin": decimal_to_str((profit_after_tax / gross_revenue * Decimal("100")) if gross_revenue > 0 else Decimal("0")),
            "costCompletenessPercent": decimal_to_str(cost_completeness),
            "rowsCount": rows_count,
            "productsCount": products_count,
        }

    def product_breakdown(self, date_from: date, date_to: date, *, sort: str = "profitAfterTax", order: str = "desc") -> list[dict[str, Any]]:
        validate_date_range(date_from, date_to)
        rows = self._query_rows(date_from, date_to).all()
        products, _ = self._product_rollup(rows, date_from, date_to)
        reverse = order.lower() != "asc"
        return sorted(products, key=lambda item: Decimal(str(item.get(sort, "0"))), reverse=reverse)

    def timeline(self, date_from: date, date_to: date, *, group_by: str) -> list[dict[str, Any]]:
        validate_date_range(date_from, date_to)
        
        if group_by == "month":
            month_from = date(date_from.year, date_from.month, 1)
            month_to = date(date_to.year, date_to.month, 1)
            summaries = (
                self._db.query(WbFinancialMonthlySummary)
                .filter(
                    WbFinancialMonthlySummary.seller_id == self._seller.id,
                    WbFinancialMonthlySummary.summary_month >= month_from,
                    WbFinancialMonthlySummary.summary_month <= month_to,
                )
                .order_by(WbFinancialMonthlySummary.summary_month.asc())
                .all()
            )
            return [
                {
                    "bucket": s.summary_month.strftime("%Y-%m"),
                    "forPay": decimal_to_str(s.for_pay),
                    "grossRevenue": decimal_to_str(s.gross_revenue),
                    "profitAfterTax": decimal_to_str(s.profit_after_tax),
                }
                for s in summaries
            ]
        elif group_by == "year":
            summaries = (
                self._db.query(WbFinancialMonthlySummary)
                .filter(
                    WbFinancialMonthlySummary.seller_id == self._seller.id,
                    WbFinancialMonthlySummary.summary_month >= date_from,
                    WbFinancialMonthlySummary.summary_month <= date_to,
                )
                .all()
            )
            buckets = defaultdict(lambda: {"forPay": Decimal("0"), "grossRevenue": Decimal("0"), "profitAfterTax": Decimal("0")})
            for s in summaries:
                key = s.summary_month.strftime("%Y")
                buckets[key]["forPay"] += s.for_pay
                buckets[key]["grossRevenue"] += s.gross_revenue
                buckets[key]["profitAfterTax"] += s.profit_after_tax
            return [
                {
                    "bucket": key,
                    "forPay": decimal_to_str(val["forPay"]),
                    "grossRevenue": decimal_to_str(val["grossRevenue"]),
                    "profitAfterTax": decimal_to_str(val["profitAfterTax"]),
                }
                for key, val in sorted(buckets.items())
            ]
        else: # default daily or weekly
            summaries = (
                self._db.query(WbFinancialDailySummary)
                .filter(
                    WbFinancialDailySummary.seller_id == self._seller.id,
                    WbFinancialDailySummary.summary_date >= date_from,
                    WbFinancialDailySummary.summary_date <= date_to,
                )
                .order_by(WbFinancialDailySummary.summary_date.asc())
                .all()
            )
            if group_by == "week":
                buckets = defaultdict(lambda: {"forPay": Decimal("0"), "grossRevenue": Decimal("0"), "profitAfterTax": Decimal("0")})
                for s in summaries:
                    monday = s.summary_date - timedelta(days=s.summary_date.weekday())
                    key = monday.isoformat()
                    buckets[key]["forPay"] += s.for_pay
                    buckets[key]["grossRevenue"] += s.gross_revenue
                    buckets[key]["profitAfterTax"] += s.profit_after_tax
                return [
                    {
                        "bucket": key,
                        "forPay": decimal_to_str(val["forPay"]),
                        "grossRevenue": decimal_to_str(val["grossRevenue"]),
                        "profitAfterTax": decimal_to_str(val["profitAfterTax"]),
                    }
                    for key, val in sorted(buckets.items())
                ]
            else:
                return [
                    {
                        "bucket": s.summary_date.isoformat(),
                        "forPay": decimal_to_str(s.for_pay),
                        "grossRevenue": decimal_to_str(s.gross_revenue),
                        "profitAfterTax": decimal_to_str(s.profit_after_tax),
                    }
                    for s in summaries
                ]

    def cost_breakdown(self, date_from: date, date_to: date) -> dict[str, Any]:
        validate_date_range(date_from, date_to)
        
        summaries = (
            self._db.query(WbFinancialDailySummary)
            .filter(
                WbFinancialDailySummary.seller_id == self._seller.id,
                WbFinancialDailySummary.summary_date >= date_from,
                WbFinancialDailySummary.summary_date <= date_to,
            )
            .all()
        )
        
        wb_costs = sum((s.wb_costs for s in summaries), Decimal("0"))
        cogs = sum((s.cogs for s in summaries), Decimal("0"))
        
        costs = (
            self._db.query(ExternalCost)
            .filter(
                ExternalCost.seller_id == self._seller.id,
                or_(
                    and_(ExternalCost.period_from.is_not(None), ExternalCost.period_to.is_not(None), ExternalCost.period_from <= date_to, ExternalCost.period_to >= date_from),
                    and_(ExternalCost.period_from.is_(None), ExternalCost.cost_date >= date_from, ExternalCost.cost_date <= date_to),
                ),
            )
            .all()
        )
        external_costs = sum((decimal_or_zero(c.amount) for c in costs), Decimal("0"))
        
        return {
            "wbCosts": decimal_to_str(wb_costs),
            "cogs": decimal_to_str(cogs),
            "externalAllocatedCosts": decimal_to_str(external_costs),
        }

    def insights(self, date_from: date, date_to: date) -> list[dict[str, Any]]:
        validate_date_range(date_from, date_to)
        products = self.product_breakdown(date_from, date_to)
        insights: list[dict[str, Any]] = []
        missing_cost_products = [item for item in products if not item["hasCostSettings"]]
        if missing_cost_products:
            insights.append(
                {
                    "type": "missing_cost_settings",
                    "level": "warning",
                    "message": f"{len(missing_cost_products)} products are missing cost settings.",
                    "affectedMetric": "costCompletenessPercent",
                    "productIds": [item["productId"] for item in missing_cost_products if item["productId"] is not None],
                    "recommendedAction": "Fill product finance settings before trusting profit analytics.",
                }
            )
        negative_profit = [item for item in products if decimal_or_zero(item["profitAfterTax"]) < 0]
        if negative_profit:
            insights.append(
                {
                    "type": "negative_profit_products",
                    "level": "danger",
                    "message": f"{len(negative_profit)} products have negative profit after tax.",
                    "affectedMetric": "profitAfterTax",
                    "productIds": [item["productId"] for item in negative_profit if item["productId"] is not None],
                    "recommendedAction": "Review cost inputs, pricing, and WB fee-heavy SKUs.",
                }
            )
        low_margin = [item for item in products if decimal_or_zero(item["grossRevenue"]) > 0 and decimal_or_zero(item["profitMargin"]) <= Decimal("5")]
        if low_margin:
            insights.append(
                {
                    "type": "low_profit_margin",
                    "level": "warning",
                    "message": f"{len(low_margin)} products have revenue but very low profit margin.",
                    "affectedMetric": "profitMargin",
                    "productIds": [item["productId"] for item in low_margin if item["productId"] is not None],
                    "recommendedAction": "Reprice or reduce variable costs for low-margin products.",
                }
            )
        return insights

    def allocation_preview(self, date_from: date, date_to: date) -> dict[str, Any]:
        validate_date_range(date_from, date_to)
        rows = self._query_rows(date_from, date_to).all()
        _, totals = self._product_rollup(rows, date_from, date_to)
        allocations = self._external_cost_allocations(date_from, date_to, totals)
        items = []
        for product_id, amount in allocations.items():
            product = self._db.get(WbProduct, product_id) if product_id else None
            items.append(
                {
                    "productId": product_id,
                    "nmId": product.nm_id if product else None,
                    "vendorCode": product.vendor_code if product else None,
                    "allocatedAmount": decimal_to_str(amount),
                }
            )
        return {"items": items}


class FinanceAiAnalysisService:
    def __init__(self, db: Session, settings: Settings, seller: Seller) -> None:
        self._db = db
        self._settings = settings
        self._seller = seller

    def analyze(self, *, date_from: date, date_to: date, group_by: str = "day") -> dict[str, Any]:
        validate_date_range(date_from, date_to)
        aggregation = FinanceAggregationService(self._db, self._seller)
        summary = aggregation.summary(date_from, date_to)
        products = aggregation.product_breakdown(date_from, date_to)
        insights = aggregation.insights(date_from, date_to)
        cost_breakdown = aggregation.cost_breakdown(date_from, date_to)

        ai_payload = {
            "summary": summary,
            "costBreakdown": cost_breakdown,
            "topProfitableProducts": products[:5],
            "topLossProducts": sorted(products, key=lambda item: Decimal(item["profitAfterTax"]))[:5],
            "missingCostProductsCount": len([item for item in products if not item["hasCostSettings"]]),
            "insights": insights,
        }

        if not self._settings.gemini_api_key:
            result = {
                "status": "gemini_not_configured",
                "analysisType": "finance_overview",
                "summary": "Gemini key is not configured. Deterministic finance analytics remain available.",
                "sections": [],
            }
        else:
            client = genai.Client(api_key=self._settings.gemini_api_key)
            prompt = (
                "Analyze this marketplace finance payload and return JSON with keys "
                "`analysisType`, `summary`, `sections`, and `actions`. "
                "Do not echo secrets. Focus on business insights.\n\n"
                f"{json.dumps(ai_payload, ensure_ascii=False)}"
            )
            response = client.models.generate_content(
                model=self._settings.gemini_model,
                contents=[prompt],
                config=types.GenerateContentConfig(response_mime_type="application/json", temperature=0.1),
            )
            try:
                result = json.loads(response.text or "{}")
            except Exception:
                result = {"analysisType": "finance_overview", "summary": response.text or "", "sections": []}

        snapshot = FinanceAnalysisSnapshot(
            seller_id=self._seller.id,
            date_from=date_from,
            date_to=date_to,
            group_by=group_by,
            summary=summary,
            product_breakdown=products,
            cost_breakdown=cost_breakdown,
            insights=insights,
            ai_analysis=result,
        )
        self._db.add(snapshot)
        self._db.commit()
        self._db.refresh(snapshot)
        return {"snapshotId": snapshot.id, "analysis": result}

    def list_snapshots(self) -> list[FinanceAnalysisSnapshot]:
        return (
            self._db.query(FinanceAnalysisSnapshot)
            .filter(FinanceAnalysisSnapshot.seller_id == self._seller.id)
            .order_by(FinanceAnalysisSnapshot.id.desc())
            .all()
        )

    async def reconciliation(self, *, date_from: date, date_to: date, client: WbFinanceClient, period: str = "daily") -> dict[str, Any]:
        validate_date_range(date_from, date_to)
        aggregation = FinanceAggregationService(self._db, self._seller)
        summary = aggregation.summary(date_from, date_to)
        report_list: list[dict[str, Any]] = []
        warning = None
        try:
            report_list = await client.get_sales_reports_list(date_from=dt_start(date_from), date_to=dt_end(date_to), period=period)
        except Exception as exc:
            warning = str(exc)
        if not report_list and warning is None:
            warning = "Finance report list is unavailable or returned no rows for the selected period."
        list_totals = defaultdict(lambda: Decimal("0"))
        for row in report_list:
            for api_key, out_key in (
                ("retailAmountSum", "retailAmountSum"),
                ("forPaySum", "forPaySum"),
                ("deliveryServiceSum", "deliveryServiceSum"),
                ("paidStorageSum", "paidStorageSum"),
                ("paidAcceptanceSum", "paidAcceptanceSum"),
                ("deductionSum", "deductionSum"),
                ("penaltySum", "penaltySum"),
                ("additionalPaymentSum", "additionalPaymentSum"),
            ):
                list_totals[out_key] += decimal_or_zero(row.get(api_key))
        return {
            "warning": warning,
            "calculatedSummary": summary,
            "reportListCount": len(report_list),
            "reportListTotals": {key: decimal_to_str(value) for key, value in list_totals.items()},
            "differences": {
                "retailAmountSum": decimal_to_str(list_totals["retailAmountSum"] - decimal_or_zero(summary["grossRevenue"])),
                "forPaySum": decimal_to_str(list_totals["forPaySum"] - decimal_or_zero(summary["forPay"])),
                "deliveryServiceSum": decimal_to_str(list_totals["deliveryServiceSum"] - decimal_or_zero(aggregation.cost_breakdown(date_from, date_to)["wbCosts"])),
            },
        }


class FinanceSystemStatusService:
    def __init__(self, db: Session, settings: Settings, seller: Seller) -> None:
        self._db = db
        self._settings = settings
        self._seller = seller

    async def build(self) -> dict[str, Any]:
        active_cooldowns = await get_active_cooldowns(seller_id=self._seller.id)
        automation_state = (
            self._db.query(SellerFinanceAutomationState)
            .filter(SellerFinanceAutomationState.seller_id == self._seller.id)
            .one_or_none()
        )
        product_state = (
            self._db.query(WbProductSyncState)
            .filter(WbProductSyncState.seller_id == self._seller.id)
            .order_by(WbProductSyncState.updated_at.desc())
            .first()
        )
        finance_states = (
            self._db.query(WbFinanceSyncState)
            .filter(WbFinanceSyncState.seller_id == self._seller.id)
            .order_by(WbFinanceSyncState.updated_at.desc())
            .all()
        )
        if automation_state and automation_state.bootstrap_status in {"queued", "running"}:
            bootstrap_finance_state = next(
                (
                    state
                    for state in finance_states
                    if state.period == "daily"
                    and state.date_from == automation_state.bootstrap_range_from
                    and state.date_to == automation_state.bootstrap_range_to
                ),
                None,
            )
            if (
                product_state
                and product_state.status == "completed"
                and bootstrap_finance_state
                and bootstrap_finance_state.status == "completed"
            ):
                automation_state.bootstrap_status = "completed"
                automation_state.bootstrap_finished_at = bootstrap_finance_state.finished_at or datetime.now(UTC)
                automation_state.bootstrap_last_error = None
                automation_state.last_successful_daily_sync_date = automation_state.bootstrap_range_to
                automation_state.last_attempted_daily_sync_date = automation_state.bootstrap_range_to
                automation_state.last_daily_status = "completed"
                automation_state.last_daily_error = None
                self._db.commit()
            elif (
                product_state and product_state.status in {"failed", "interrupted"}
            ) or (bootstrap_finance_state and bootstrap_finance_state.status == "failed"):
                automation_state.bootstrap_status = "failed"
                automation_state.bootstrap_finished_at = datetime.now(UTC)
                automation_state.bootstrap_last_error = (
                    (product_state.last_error if product_state and product_state.status in {"failed", "interrupted"} else None)
                    or (bootstrap_finance_state.last_error if bootstrap_finance_state else None)
                    or "Failed during bootstrap sync."
                )[:1000]
                self._db.commit()
        last_successful_finance = next((state.finished_at for state in finance_states if state.status == "completed" and state.finished_at), None)
        failed_states = [state for state in finance_states if state.status in {"failed", "rate_limited"} and state.finished_at]
        product_failed = product_state if product_state and product_state.status in {"failed", "rate_limited"} and product_state.finished_at else None
        last_failed_candidates = [
            (state.finished_at, state.last_error) for state in failed_states
        ]
        if product_failed:
            last_failed_candidates.append((product_failed.finished_at, product_failed.last_error))
        last_failed_candidates = [item for item in last_failed_candidates if item[0] is not None]
        last_failed_candidates.sort(key=lambda item: item[0], reverse=True)
        last_failed_at = last_failed_candidates[0][0] if last_failed_candidates else None
        last_failed_error = last_failed_candidates[0][1] if last_failed_candidates else None

        missing_settings_count = (
            self._db.query(func.count(WbProduct.id))
            .filter(
                WbProduct.seller_id == self._seller.id,
                ~WbProduct.id.in_(
                    self._db.query(ProductFinanceSetting.product_id).filter(ProductFinanceSetting.seller_id == self._seller.id)
                ),
            )
            .scalar()
            or 0
        )
        unmapped_rows_count = (
            self._db.query(func.count(WbFinanceReportRow.id))
            .filter(WbFinanceReportRow.seller_id == self._seller.id, WbFinanceReportRow.product_id.is_(None))
            .scalar()
            or 0
        )
        next_run_at = self._next_scheduled_run_at(automation_state.timezone if automation_state else self._settings.finance_auto_sync_timezone)

        return {
            "contentApi": self._availability(active_cooldowns, "content"),
            "financeApi": self._availability(active_cooldowns, "finance"),
            "commonApi": self._availability(active_cooldowns, "common"),
            "sellerInfoApi": self._availability(active_cooldowns, "common", endpoint="/api/v1/seller-info"),
            "activeCooldowns": active_cooldowns,
            "lastSuccessfulProductSyncAt": product_state.finished_at.isoformat() if product_state and product_state.status == "completed" and product_state.finished_at else None,
            "lastSuccessfulFinanceSyncAt": last_successful_finance.isoformat() if last_successful_finance else None,
            "lastFailedSyncAt": last_failed_at.isoformat() if last_failed_at else None,
            "lastFailedSyncError": last_failed_error[:300] if last_failed_error else None,
            "geminiConfigured": bool(self._settings.gemini_api_key),
            "hasProductsMissingFinanceSettings": missing_settings_count > 0,
            "missingFinanceSettingsCount": int(missing_settings_count),
            "hasUnmappedFinanceRows": unmapped_rows_count > 0,
            "unmappedFinanceRowsCount": int(unmapped_rows_count),
            "automationTimezone": automation_state.timezone if automation_state else None,
            "bootstrapStatus": automation_state.bootstrap_status if automation_state else None,
            "bootstrapRangeFrom": automation_state.bootstrap_range_from.isoformat() if automation_state and automation_state.bootstrap_range_from else None,
            "bootstrapRangeTo": automation_state.bootstrap_range_to.isoformat() if automation_state and automation_state.bootstrap_range_to else None,
            "bootstrapFinishedAt": automation_state.bootstrap_finished_at.isoformat() if automation_state and automation_state.bootstrap_finished_at else None,
            "lastSuccessfulDailySyncDate": automation_state.last_successful_daily_sync_date.isoformat() if automation_state and automation_state.last_successful_daily_sync_date else None,
            "lastDailySyncStatus": automation_state.last_daily_status if automation_state else None,
            "lastDailySyncError": automation_state.last_daily_error[:300] if automation_state and automation_state.last_daily_error else None,
            "nextScheduledRunAt": next_run_at.isoformat() if next_run_at else None,
        }

    @staticmethod
    def _availability(active_cooldowns: list[dict[str, Any]], category: str, endpoint: str | None = None) -> dict[str, Any]:
        items = [item for item in active_cooldowns if item["category"] == category and (endpoint is None or item["endpoint"] == endpoint)]
        return {
            "available": len(items) == 0,
            "inCooldown": len(items) > 0,
            "activeCooldownCount": len(items),
            "cooldowns": items,
        }

    @staticmethod
    def _next_scheduled_run_at(timezone_name: str) -> datetime | None:
        try:
            tz = ZoneInfo(timezone_name)
        except Exception:
            return None
        local_now = datetime.now(UTC).astimezone(tz)
        next_local = datetime.combine(local_now.date() + timedelta(days=1), time.min, tzinfo=tz)
        return next_local.astimezone(UTC)
